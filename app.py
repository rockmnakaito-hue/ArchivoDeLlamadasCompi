import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import random

MANAGERS = ["Rockman", "Diego", "Chane", "Parka", "Jose", "Greeg", "Gledys", "Salem", "Diana", "Josue"]

def agregar_menus_y_formato(ws, max_row):
    valores_si_no = ["SI", "NO"]
    dv_si_no = DataValidation(type="list", formula1=f'"{",".join(valores_si_no)}"', allow_blank=True)
    ws.add_data_validation(dv_si_no)
    dv_si_no.add(f"F2:F{max_row}")
    dv_si_no.add(f"G2:G{max_row}")

    color_map_si_no = {
        "SI": "00FF00",
        "NO": "FF0000"
    }
    for col in [6, 7]:  # F=6, G=7
        for valor, color in color_map_si_no.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            rule = FormulaRule(formula=[f'${chr(64+col)}2="{valor}"'], fill=fill)
            ws.conditional_formatting.add(f'{chr(64+col)}2:{chr(64+col)}{max_row}', rule)

    valores_valoracion = ["Normal", "Buena", "Mejorable", "La cago"]
    dv_valoracion = DataValidation(type="list", formula1=f'"{",".join(valores_valoracion)}"', allow_blank=True)
    ws.add_data_validation(dv_valoracion)
    dv_valoracion.add(f"I2:I{max_row}")

    color_map_valoracion = {
        "Normal": "FFFF00",
        "Buena": "00FF00",
        "Mejorable": "FFC7CE",
        "La cago": "9C0006"
    }
    for valor, color in color_map_valoracion.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = FormulaRule(formula=[f'$I2="{valor}"'], fill=fill)
        ws.conditional_formatting.add(f'I2:I{max_row}', rule)

def repartir_llamadas_balanceado(df, minutos_objetivo, managers, margen_extra=4):
    df = df.copy().reset_index(drop=True)
    total_minutos = df['Minutos'].sum()
    total_requerido = minutos_objetivo * len(managers)

    if total_minutos < total_requerido:
        st.warning(f"ADVERTENCIA: No hay suficientes minutos totales ({total_minutos}) para cubrir {total_requerido}.")
        return None

    df = df.sample(frac=1, random_state=random.randint(0,10000)).reset_index(drop=True)

    asignaciones = {m: [] for m in managers}
    minutos_por_manager = {m: 0 for m in managers}
    usados_indices = set()

    for manager in managers:
        suma = 0
        for idx, row in df.iterrows():
            if idx in usados_indices:
                continue
            if suma >= minutos_objetivo and suma <= minutos_objetivo + margen_extra:
                break
            if suma + row['Minutos'] > minutos_objetivo + margen_extra:
                if suma >= minutos_objetivo:
                    break
                else:
                    continue
            asignaciones[manager].append(row)
            minutos_por_manager[manager] += row['Minutos']
            suma += row['Minutos']
            usados_indices.add(idx)

    asignaciones_df = {}
    for m in managers:
        asignaciones_df[m] = pd.DataFrame(asignaciones[m])

    return asignaciones_df, usados_indices

def procesar_csv(df, minutos_objetivo=None, min_minutos=3, max_minutos=20):
    cols_to_drop = [
        'id', 'fromNumber',
        'toNumber', 'toName', 'viaNumber', 'dateAnswered', 'dateFinished'
    ]
    df = df.drop(columns=cols_to_drop, errors='ignore')

    if 'fromName' in df.columns:
        df = df.rename(columns={'fromName': 'Agente'})
    else:
        df.insert(1, "Agente", "")

    df['callDuration'] = df['callDuration'].str.replace('segs', '', regex=False).str.strip()

    df['is_number'] = df['callDuration'].apply(lambda x: str(x).replace(" ", "").isdigit())
    df = df[df['is_number'] == False].copy()
    df.drop(columns=['is_number'], inplace=True)

    dur_split = df['callDuration'].str.split(' ', expand=True)
    df['callDuration'] = dur_split[0]

    df['callDuration'] = pd.to_numeric(df['callDuration'], errors='coerce')
    df = df[(df['callDuration'] >= min_minutos) & (df['callDuration'] <= max_minutos)]

    df = df.rename(columns={'callDuration': 'Minutos'})

    df['Se presentó?'] = ""
    df['Nota'] = ""

    df['Enlaces'] = ""
    df['Valoración'] = ""
    df['Nota2'] = ""
    df['LinkBase'] = "https://compinche.ladesk.com/agent/index.php?rnd=8146#Conversation;id="
    df['Enlaces'] = df['LinkBase'] + df['ticketId']
    df.drop(columns=['LinkBase'], inplace=True)

    asignaciones_df = None
    usados_indices = set()
    if minutos_objetivo is not None:
        resultado = repartir_llamadas_balanceado(df, minutos_objetivo, MANAGERS)
        if resultado is not None:
            asignaciones_df, usados_indices = resultado

            # Solo eliminar índices que están presentes
            indices_a_eliminar = [i for i in usados_indices if i in df.index]
            df = df.drop(index=indices_a_eliminar).reset_index(drop=True)

    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    ws.title = "Base de Datos"

    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)

    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    col_f = 6
    col_g = 7
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=col_f).fill = fill_white
        ws.cell(row=row, column=col_f).alignment = align_center
        ws.cell(row=row, column=col_g).fill = fill_white
        ws.cell(row=row, column=col_g).alignment = align_center

    agregar_menus_y_formato(ws, max_row)

    if asignaciones_df is not None:
        for manager, df_man in asignaciones_df.items():
            ws_man = wb.create_sheet(title=manager)
            if df_man.empty:
                for c_idx, col_name in enumerate(df.columns, start=1):
                    ws_man.cell(row=1, column=c_idx, value=col_name).font = Font(bold=True, size=12)
            else:
                for c_idx, col_name in enumerate(df_man.columns, start=1):
                    ws_man.cell(row=1, column=c_idx, value=col_name).font = Font(bold=True, size=12)
                for r_idx, row in enumerate(df_man.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws_man.cell(row=r_idx, column=c_idx, value=value)

            max_row_man = ws_man.max_row
            agregar_menus_y_formato(ws_man, max_row_man)

    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)

    return output_final

st.title("Procesar CSV a Excel Formateado")

uploaded_file = st.file_uploader("Selecciona archivo CSV", type=["csv"])

minutos = st.number_input("¿Cuántos minutos debe escuchar cada manager?", min_value=1, value=25)
min_llamada = st.number_input("Llamadas mínimo (minutos)", min_value=0, value=3)
max_llamada = st.number_input("Llamadas máximo (minutos)", min_value=0, value=20)

if uploaded_file is not None:
    try:
        df = pd.read_csv(uploaded_file)
        archivo_excel = procesar_csv(df, minutos_objetivo=minutos, min_minutos=min_llamada, max_minutos=max_llamada)
        if archivo_excel is None:
            st.warning("No hay suficientes minutos totales para cubrir a todos los managers. Se generará sólo la hoja base.")
            archivo_excel = procesar_csv(df, minutos_objetivo=None)

        st.download_button(
            label="Descargar archivo Excel",
            data=archivo_excel,
            file_name="llamadas_procesadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
